from openpyxl import load_workbook
from datetime import datetime, timedelta

from sender.models import NicOnSurname, BotSetting


class BaseMessage:
    def __init__(self, dst, data_row, target_date: datetime):
        """
        :param dst: имя файла
        :param data_row: строка, в которой ищем число
        """

        self.wb = load_workbook(dst)
        self.data_row = data_row

        self.target_date = target_date

        # ищем в книге имя листа, в котором содержится текущий месяц
        month = self.get_name_month(self.target_date.month)
        sheet_name = self.get_sheet_name(month, self.wb.sheetnames)
        self.ws = self.wb[sheet_name]

        # буква столбца с ФИО
        self.employer_col = BotSetting.objects.get(pk=1).col_number_for_nic.upper()

    @staticmethod
    def get_name_month(month) -> str | bool:

        month_list = [
            'ноль',
            'январь', 'февраль', 'март',
            'апрель', 'май', 'июнь',
            'июль', 'август', 'сентябрь',
            'октябрь', 'ноябрь', 'декабрь']

        if 0 < month <= 12:
            return month_list[month]

        return False

    @staticmethod
    def get_sheet_name(keyword: str, sheets_names: list) -> str | bool:

        for name in sheets_names:
            if keyword in name.lower():
                return name

        return False

    @staticmethod
    def get_column_letter_by_day(row, day) -> str | bool:

        for cell in row:
            if str(day) in str(cell.value):
                return cell.column_letter

        return False


class DutyMessage(BaseMessage):
    """
    Формирует сообщение для дежурных по чатам
    """

    def __init__(self, dst, data_row, green, red):
        """
        :param green: rgb для зеленого цвета
        :param red: rgb для красного
        """
        super().__init__(dst, data_row, datetime.now())

        # цвета в rgb для поиска
        self.green = green
        self.red = red

    def get_one_message(self, color: list) -> str | None:

        row = self.ws[self.data_row]

        # получаем букву столбца по дню
        column_letter = self.get_column_letter_by_day(row, self.target_date.day)

        # получаем номер ячейки по цвету текста
        chell_number = self.get_cell_row_by_font_color(color, self.ws[column_letter])

        if not chell_number:
            return None

        if not self.employer_col:
            employer = 'B' + str(chell_number)
        else:
            employer = self.employer_col + str(chell_number)
        employer = str(self.ws[employer].value).split()

        # Если ник есть в таблице, используем его
        nic = NicOnSurname.objects.filter(surname__istartswith=employer[0]).first()
        if nic:
            return f'{nic.nic}'

        return f'{" ".join(employer[:2])}'

    def build(self):

        green_message = self.get_one_message([self.green])

        if green_message:
            green_message += ', сегодня ты дежурный по #ncc-call-tech'
        else:
            green_message = 'по #ncc-call-tech сегодня дежурных нет'

        # TODO: нормально обработать цвета, пока костыль
        red_message = self.get_one_message([self.red, 2])

        if red_message:
            red_message += ', сегодня ты дежурный по #DomainsDuty'
        else:
            red_message = 'по #DomainsDuty сегодня дежурных нет'

        return green_message, red_message

    @staticmethod
    def get_cell_row_by_font_color(color: list, ws_col) -> int | None:
        for cell in ws_col:
            if cell.font.color:
                if cell.font.color.value in color:
                    if cell.value:
                        return cell.row
        return None


class HolidayMessage(BaseMessage):

    def __init__(self, dst, data_row, interval, blue):
        """
        :param blue: цвет ячейки неоформленного отпуска
        """

        target_date = datetime.now() + timedelta(days=interval)
        super().__init__(dst, data_row, target_date)

        self.blue = blue

    def get_employers(self, color: str) -> list[str] | bool:

        row = self.ws[self.data_row]

        # получаем букву столбца по дню
        column_letter = self.get_column_letter_by_day(row, self.target_date.day)

        # получаем номера ячеек по цвету фона
        chell_numbers = self.get_cell_row_by_color(color, self.ws[column_letter])

        if not chell_numbers:
            return False

        employers = []

        for chell_number in chell_numbers:

            if not self.employer_col:
                employer = 'B' + str(chell_number)
            else:
                employer = self.employer_col + str(chell_number)

            employer = str(self.ws[employer].value).split()

            # Если ник есть в таблице, используем его
            nic = NicOnSurname.objects.filter(surname__istartswith=employer[0]).first()
            if nic:
                employers.append(str(nic.nic))
                continue

            employers.append(" ".join(employer[:2]))

        return employers

    def build(self) -> str | None:
        employers = self.get_employers(self.blue)
        if employers:
            message = ', '.join(employers)
            message += ' - не отправлено заявление на отпуск'
            message += ', вот инструкция как это сделать https://confluence.runity.ru/x/crJtBQ'
            return message
        return None

    @staticmethod
    def get_cell_row_by_color(color: str, ws_col) -> list | list[int]:
        cell_rows = []

        for cell in ws_col:
            if cell.fill.fgColor.rgb == color:
                cell_rows.append(cell.row)

        return cell_rows
